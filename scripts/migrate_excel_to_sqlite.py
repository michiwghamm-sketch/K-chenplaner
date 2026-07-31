from __future__ import annotations

import argparse
import json
import sys
from dataclasses import asdict, dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select

PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from app.config import AppConfig
from app.db import initialize_database, session_scope
from app.models import (
    CampYear,
    ImportIssue,
    ImportRun,
    Ingredient,
    IngredientPrice,
    MealPlanEntry,
    Recipe,
    RecipeComponent,
    RecipeFeedback,
    RecipeIngredient,
    RecipeStep,
    ShoppingList,
    ShoppingListItem,
)
from app.services.data_cleanup_service import NON_INGREDIENT_EXACT_NAMES, is_non_ingredient_name
from app.utils.normalization import normalize_name
from app.utils.units import normalize_unit, parse_decimal
from scripts.inspect_excel import find_excel_file


RECIPE_EXCLUDED_SHEETS = {
    "Preisliste 2024",
    "Einkaufsliste 2024",
    "Preisliste",
    "Einkaufsliste 2025",
    "Rezepte Übersicht",
    "Vorlage",
    "Preisquellen & Pflege",
    "Planung 2026",
    "Rezept Feedback",
    "Einkaufsplanung",
}


@dataclass
class ImportCounters:
    ingredients: int = 0
    ingredient_prices: int = 0
    recipes: int = 0
    recipe_components: int = 0
    recipe_ingredients: int = 0
    recipe_steps: int = 0
    camp_years: int = 0
    meal_plan_entries: int = 0
    recipe_feedback: int = 0
    shopping_lists: int = 0
    shopping_list_items: int = 0
    import_issues: int = 0


@dataclass
class ImportIssueRecord:
    severity: str
    sheet_name: str | None
    cell_reference: str | None
    message: str
    raw_value: str | None = None


def is_recipe_sheet(sheet_name: str) -> bool:
    return sheet_name not in RECIPE_EXCLUDED_SHEETS


def worksheet_rows(ws: Worksheet, *, min_col: int = 1, max_col: int | None = None) -> Iterable[list[object]]:
    limit = max_col or ws.max_column
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=min_col, max_col=limit, values_only=True):
        yield list(row)


def find_recipe_portions(ws: Worksheet) -> int | None:
    value = ws["H6"].value
    return int(value) if isinstance(value, (int, float)) else None


def extract_recipe_instructions(ws: Worksheet) -> str | None:
    lines: list[str] = []
    in_steps = False
    for row in range(1, ws.max_row + 1):
        title = ws.cell(row=row, column=1).value
        action = ws.cell(row=row, column=2).value
        # Spalte G ("Ungefaehre Dauer in min:"), nicht Spalte C - siehe extract_recipe_steps().
        duration = ws.cell(row=row, column=7).value
        if title == "Schritte:":
            in_steps = True
            continue
        if not in_steps:
            continue
        if title == "Gesamtdauer:":
            break
        if title or action:
            part = f"{title or ''}: {action or ''}".strip(": ").strip()
            if duration not in (None, ""):
                part = f"{part} ({duration} min)"
            lines.append(part)
    return "\n".join(lines) if lines else None


def extract_recipe_steps(ws: Worksheet) -> list[dict]:
    """Liest die Arbeitsschritte (Titel/Anweisung/Dauer) strukturiert aus dem 'Schritte:'-Block.

    Layout im Workbook: Spalte A = Titel, Spalte B = Anweisung, Spalte G = Dauer in Minuten.
    Die Zeilennummer dient als stabiler sort_order/Idempotenz-Schluessel, analog zu den
    Rezeptzutaten (siehe import_recipe_sheet).
    """
    steps: list[dict] = []
    in_steps = False
    for row in range(1, ws.max_row + 1):
        title = ws.cell(row=row, column=1).value
        action = ws.cell(row=row, column=2).value
        duration = ws.cell(row=row, column=7).value
        if title == "Schritte:":
            in_steps = True
            continue
        if not in_steps:
            continue
        if title == "Gesamtdauer:":
            break
        if not title and not action:
            continue
        steps.append(
            {
                "row": row,
                "title": str(title).strip() if title else None,
                "description": str(action).strip() if action else None,
                "duration_minutes": int(duration) if isinstance(duration, (int, float)) else None,
            }
        )
    return steps


def recipe_category_from_name(name: str) -> str | None:
    lowered = normalize_name(name)
    if "fruhstuck" in lowered:
        return "Frühstück"
    if "salat" in lowered:
        return "Salat"
    return None


def recipe_meal_type_from_name(name: str) -> str | None:
    lowered = normalize_name(name)
    if "fruhstuck" in lowered:
        return "Frühstück"
    return "Gericht"


def resolve_or_create_ingredient(
    session, cache: dict[str, Ingredient], name: str, unit: str | None, *, set_default_unit: bool = True
) -> tuple[Ingredient, bool]:
    """Findet oder legt eine Zutat an. 'unit' setzt nur dann die Standardeinheit, wenn
    set_default_unit=True - die Preisliste liefert in Spalte 3 eine Preis-Einheit (z. B. 'kg' bei
    '2,50 EUR pro kg'), keine Mess-Einheit fuer Rezeptmengen, und darf die Standardeinheit deshalb
    nicht setzen (siehe scripts/cleanup_units.py fuer die Bereinigung der historischen Vermischung)."""
    normalized = normalize_name(name)
    if normalized in cache:
        ingredient = cache[normalized]
        if set_default_unit and unit and not ingredient.default_unit:
            ingredient.default_unit = unit
        return ingredient, False

    ingredient = session.execute(
        select(Ingredient).where(Ingredient.normalized_name == normalized)
    ).scalar_one_or_none()
    created = False
    if ingredient is None:
        ingredient = Ingredient(
            name=name.strip(),
            normalized_name=normalized,
            default_unit=unit if set_default_unit else None,
        )
        session.add(ingredient)
        created = True
    elif set_default_unit and unit and not ingredient.default_unit:
        ingredient.default_unit = unit

    cache[normalized] = ingredient
    return ingredient, created


def should_skip_price_list_row(name: object) -> bool:
    if not isinstance(name, str):
        return not bool(name)
    return is_non_ingredient_name(name)


def should_skip_shopping_row(name: object, unit: str | None) -> bool:
    if not isinstance(name, str):
        return not bool(name)
    if is_non_ingredient_name(name):
        return True
    return (unit or "").strip().lower() == "portionen"


def import_price_list(ws: Worksheet, session, ingredient_cache: dict[str, Ingredient], counters: ImportCounters, issues: list[ImportIssueRecord]) -> None:
    header = [ws.cell(row=1, column=col).value for col in range(1, 8)]
    if not header or header[0] != "Zutat":
        issues.append(ImportIssueRecord("warning", ws.title, "A1", "Preisliste hat unerwarteten Header.", str(header)))
        return

    year = None
    for token in ws.title.split():
        if token.isdigit():
            year = int(token)

    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        price = ws.cell(row=row, column=2).value
        unit = ws.cell(row=row, column=3).value
        if should_skip_price_list_row(name):
            continue

        unit_text = normalize_unit(unit)
        ingredient, created = resolve_or_create_ingredient(
            session, ingredient_cache, str(name), unit_text, set_default_unit=False
        )
        if created:
            counters.ingredients += 1

        decimal_price = parse_decimal(price)
        if decimal_price is None:
            issues.append(ImportIssueRecord("warning", ws.title, f"B{row}", "Preis konnte nicht gelesen werden.", str(price)))
            continue

        source = ws.cell(row=row, column=5).value
        notes = ws.cell(row=row, column=7).value
        source_text = str(source) if source else ws.title

        # Ohne diesen Check legt jeder erneute Import-Lauf dieselbe Preiszeile ein weiteres
        # Mal an (die Preisliste hat - anders als die Rezeptblaetter - keine Zeilennummer,
        # an der ein spaeterer Lauf eine schon importierte Zeile wiedererkennen koennte).
        has_matching_price = any(
            price.price_per_unit == decimal_price
            and price.unit == (unit_text or "")
            and price.year == year
            and price.source == source_text
            for price in ingredient.prices
        )
        if has_matching_price:
            continue

        session.add(
            IngredientPrice(
                ingredient=ingredient,
                price_per_unit=decimal_price,
                unit=unit_text or "",
                source=source_text,
                year=year,
                notes=str(notes) if notes else None,
            )
        )
        counters.ingredient_prices += 1


def find_totals_row(ws: Worksheet, *, start_row: int = 9) -> int:
    """Zeilenindex von 'Gesamtkosten:' in Spalte A - die Ingredienzliste endet davor."""
    for row in range(start_row, ws.max_row + 1):
        value = ws.cell(row=row, column=1).value
        if isinstance(value, str) and "gesamtkosten" in value.lower():
            return row
    return ws.max_row + 1


NON_COMPONENT_LABELS = {"preis pro portion", "gesamtkosten"}


def detect_recipe_components(ws: Worksheet, recipe_name: str, end_row: int) -> list[tuple[int, int, str]]:
    """Liest Teilstueck-Gruppen aus den ueber mehrere Zeilen zusammengefassten Zellen in Spalte G/H.

    Jedes Teilstueck (z. B. 'Soße', 'Kartoffelbrei') ist im Excel als eine ueber mehrere
    Zeilen verbundene Zelle rechts neben der Zutatentabelle formatiert; die Zeilenspanne der
    Verbindung entspricht genau den Zutatenzeilen, die zu diesem Teilstueck gehoeren.
    """
    components: list[tuple[int, int, str]] = []
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        if min_col not in (7, 8) or not (8 <= min_row < end_row) or max_row <= min_row:
            continue
        value = ws.cell(row=min_row, column=min_col).value
        if value is None or isinstance(value, (int, float)):
            continue
        label = str(value).strip()
        if not label or label.lower() in NON_COMPONENT_LABELS:
            continue
        components.append((min_row, min(max_row, end_row - 1), label))
    components.sort()

    if len(components) == 1:
        start, stop, label = components[0]
        # Leerzeichen ignorieren: Labels wiederholen den Rezeptnamen manchmal ohne
        # Leerzeichen (z. B. "Gemüsenudeln" bei Rezept "Gemüse Nudeln").
        label_key = normalize_name(label).replace(" ", "")
        recipe_key = normalize_name(recipe_name).replace(" ", "")
        if start <= 9 and stop >= end_row - 1 and label_key == recipe_key:
            # Einzelnes Etikett, das den kompletten Zutatenblock ueberspannt und nur den
            # Rezeptnamen wiederholt - kein echtes Teilstueck, sondern nur die Kopfzeile.
            return []
    return components


def component_label_for_row(components: list[tuple[int, int, str]], row: int) -> str | None:
    for start, stop, label in components:
        if start <= row <= stop:
            return label
    return None


def build_known_units_map(values_wb) -> dict[str, str]:
    """Sammelt je Zutat die erste bekannte Einheit aus Preislisten und Rezeptblaettern.

    Dient als Fallback fuer Rezeptzeilen, die eine Menge aber keine Einheit angeben
    (z. B. Gewuerze wie 'Paprika Pulver geraeuchert'), damit solche Zeilen nicht mehr
    stillschweigend uebersprungen werden.
    """
    known: dict[str, str] = {}

    for sheet_name in ("Preisliste", "Preisliste 2024"):
        if sheet_name not in values_wb.sheetnames:
            continue
        ws = values_wb[sheet_name]
        for row in range(2, ws.max_row + 1):
            name = ws.cell(row=row, column=1).value
            normalized_unit = normalize_unit(ws.cell(row=row, column=3).value)
            if isinstance(name, str) and normalized_unit:
                known.setdefault(normalize_name(name), normalized_unit)

    for sheet_name in values_wb.sheetnames:
        if not is_recipe_sheet(sheet_name):
            continue
        ws = values_wb[sheet_name]
        end_row = find_totals_row(ws)
        for row in range(9, end_row):
            name = ws.cell(row=row, column=1).value
            normalized_unit = normalize_unit(ws.cell(row=row, column=3).value)
            if isinstance(name, str) and normalized_unit:
                known.setdefault(normalize_name(name), normalized_unit)

    return known


def import_recipe_sheet(values_ws: Worksheet, formula_ws: Worksheet, session, ingredient_cache: dict[str, Ingredient], recipe_cache: dict[str, Recipe], known_units: dict[str, str], counters: ImportCounters, issues: list[ImportIssueRecord]) -> None:
    recipe_name = str(values_ws["E2"].value or formula_ws["E2"].value or values_ws.title).strip()
    normalized_recipe_name = normalize_name(recipe_name)
    recipe = session.execute(select(Recipe).where(Recipe.normalized_name == normalized_recipe_name)).scalar_one_or_none()
    if recipe is None:
        recipe = Recipe(
            name=recipe_name,
            normalized_name=normalized_recipe_name,
            category=recipe_category_from_name(recipe_name),
            meal_type=recipe_meal_type_from_name(recipe_name),
            default_portions=find_recipe_portions(values_ws),
            instructions=extract_recipe_instructions(formula_ws),
        )
        session.add(recipe)
        counters.recipes += 1
    else:
        recipe.instructions = recipe.instructions or extract_recipe_instructions(formula_ws)
        recipe.default_portions = recipe.default_portions or find_recipe_portions(values_ws)

    recipe_cache[normalized_recipe_name] = recipe
    recipe_cache[normalize_name(values_ws.title)] = recipe
    session.flush()

    end_row = find_totals_row(values_ws)
    detected_components = detect_recipe_components(values_ws, recipe_name, end_row)

    components_by_key: dict[str, RecipeComponent] = {
        normalize_name(component.name): component for component in recipe.components
    }
    for _start, _stop, label in detected_components:
        key = normalize_name(label)
        if key in components_by_key:
            continue
        component = RecipeComponent(
            recipe=recipe,
            name=label.strip(),
            sort_order=max((c.sort_order for c in recipe.components), default=0) + 1,
        )
        session.add(component)
        session.flush()
        components_by_key[key] = component
        counters.recipe_components += 1

    for row in range(9, end_row):
        ingredient_name = values_ws.cell(row=row, column=1).value
        quantity = values_ws.cell(row=row, column=2).value
        unit = values_ws.cell(row=row, column=3).value
        price_value = values_ws.cell(row=row, column=5).value

        if ingredient_name == "Schritte:":
            break
        if ingredient_name in (None, "", "Zutaten:", "Gesamtdauer:"):
            continue
        if not isinstance(ingredient_name, str):
            continue
        if ingredient_name.lower().startswith("zubereitung"):
            continue

        optional = False
        note_text = None
        decimal_quantity = parse_decimal(quantity)
        if decimal_quantity is None:
            if quantity not in (None, ""):
                # Ein Wert stand da, konnte aber nicht als Zahl gelesen werden - das ist ein
                # echtes Datenproblem, kein "nach Geschmack"-Fall. Zeile ueberspringen.
                issues.append(
                    ImportIssueRecord(
                        "warning", values_ws.title, f"B{row}",
                        "Menge konnte nicht gelesen werden.", f"{ingredient_name}|{quantity}",
                    )
                )
                continue
            decimal_quantity = Decimal("0")
            optional = True
            note_text = "Menge nicht in Excel angegeben (nach Geschmack)"
            issues.append(
                ImportIssueRecord(
                    "info", values_ws.title, f"A{row}",
                    "Keine Menge angegeben - als optionale Zutat (nach Geschmack) importiert.",
                    str(ingredient_name),
                )
            )

        normalized_unit = normalize_unit(unit)
        if normalized_unit is None:
            fallback_unit = known_units.get(normalize_name(ingredient_name))
            normalized_unit = fallback_unit or "kg"
            issues.append(
                ImportIssueRecord(
                    "info", values_ws.title, f"C{row}",
                    f"Keine Einheit angegeben - '{normalized_unit}' aus anderer Fundstelle uebernommen."
                    if fallback_unit
                    else f"Keine Einheit angegeben - Standardeinheit '{normalized_unit}' angenommen.",
                    str(ingredient_name),
                )
            )

        ingredient, created = resolve_or_create_ingredient(session, ingredient_cache, ingredient_name, normalized_unit)
        if created:
            counters.ingredients += 1

        component_label = component_label_for_row(detected_components, row)
        component = components_by_key.get(normalize_name(component_label)) if component_label else None

        existing = next(
            (
                item
                for item in recipe.ingredients
                if item.ingredient.normalized_name == ingredient.normalized_name and item.sort_order == row
            ),
            None,
        )
        if existing is not None:
            if component is not None and existing.component_id != component.id:
                existing.component_id = component.id
            continue

        recipe.ingredients.append(
            RecipeIngredient(
                ingredient=ingredient,
                component=component,
                quantity=decimal_quantity,
                unit=normalized_unit,
                price_unit=normalize_unit(values_ws.cell(row=row, column=3).value) or normalized_unit,
                optional=optional,
                sort_order=row,
                notes=note_text,
            )
        )
        counters.recipe_ingredients += 1

        decimal_price = parse_decimal(price_value)
        if decimal_price is not None:
            has_matching_price = any(
                price.price_per_unit == decimal_price and price.unit == normalized_unit
                for price in ingredient.prices
            )
            if not has_matching_price:
                session.add(
                    IngredientPrice(
                        ingredient=ingredient,
                        price_per_unit=decimal_price,
                        unit=normalized_unit,
                        source=values_ws.title,
                    )
                )
                counters.ingredient_prices += 1

    # Bestehende Zutaten aus frueheren Importlaeufen (noch ohne Teilstueck-Zuordnung) nachtraeglich einsortieren.
    if detected_components:
        for item in recipe.ingredients:
            if item.component_id is not None:
                continue
            label = component_label_for_row(detected_components, item.sort_order)
            if not label:
                continue
            component = components_by_key.get(normalize_name(label))
            if component is not None:
                item.component_id = component.id

    existing_step_rows = {step.sort_order for step in recipe.steps}
    for step_data in extract_recipe_steps(formula_ws):
        if step_data["row"] in existing_step_rows:
            continue
        session.add(
            RecipeStep(
                recipe=recipe,
                sort_order=step_data["row"],
                title=step_data["title"],
                description=step_data["description"],
                duration_minutes=step_data["duration_minutes"],
            )
        )
        counters.recipe_steps += 1


def parse_excel_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def import_planning_sheet(ws: Worksheet, session, recipe_cache: dict[str, Recipe], counters: ImportCounters, issues: list[ImportIssueRecord]) -> CampYear | None:
    year = ws["A4"].value
    if not isinstance(year, int):
        issues.append(ImportIssueRecord("warning", ws.title, "A4", "Planungsblatt ohne gültiges Jahr.", str(year)))
        return None

    camp_year = session.execute(select(CampYear).where(CampYear.year == year)).scalar_one_or_none()
    if camp_year is None:
        camp_year = CampYear(
            year=year,
            name=f"Zeltlager {year}",
            start_date=parse_excel_date(ws["B4"].value),
            end_date=parse_excel_date(ws["C4"].value),
            notes=str(ws["K4"].value) if ws["K4"].value else None,
        )
        session.add(camp_year)
        counters.camp_years += 1

    for row in range(7, ws.max_row + 1):
        meal_type = ws.cell(row=row, column=3).value
        recipe_name = ws.cell(row=row, column=4).value
        if meal_type in (None, "") and recipe_name in (None, ""):
            continue

        recipe = None
        if recipe_name:
            normalized = normalize_name(str(recipe_name))
            recipe = recipe_cache.get(normalized) or session.execute(
                select(Recipe).where(Recipe.normalized_name == normalized)
            ).scalar_one_or_none()
            if recipe is None:
                issues.append(ImportIssueRecord("warning", ws.title, f"D{row}", "Geplantes Rezept wurde nicht gefunden.", str(recipe_name)))

        camp_year.meal_plan_entries.append(
            MealPlanEntry(
                meal_date=parse_excel_date(ws.cell(row=row, column=1).value),
                weekday=str(ws.cell(row=row, column=2).value) if ws.cell(row=row, column=2).value else None,
                meal_type=str(meal_type) if meal_type else None,
                recipe=recipe,
                planned_portions=int(ws.cell(row=row, column=5).value) if isinstance(ws.cell(row=row, column=5).value, (int, float)) else None,
                target_group=str(ws.cell(row=row, column=6).value) if ws.cell(row=row, column=6).value else None,
                shopping_date=parse_excel_date(ws.cell(row=row, column=7).value),
                shopping_group=str(ws.cell(row=row, column=8).value) if ws.cell(row=row, column=8).value else None,
                status=str(ws.cell(row=row, column=9).value) if ws.cell(row=row, column=9).value else None,
                notes=str(ws.cell(row=row, column=11).value) if ws.cell(row=row, column=11).value else None,
            )
        )
        counters.meal_plan_entries += 1

    return camp_year


def import_feedback_sheet(ws: Worksheet, session, recipe_cache: dict[str, Recipe], camp_year_cache: dict[int, CampYear], counters: ImportCounters, issues: list[ImportIssueRecord]) -> None:
    for row in range(4, ws.max_row + 1):
        year = ws.cell(row=row, column=1).value
        recipe_name = ws.cell(row=row, column=2).value
        if not year or not recipe_name:
            continue

        camp_year = camp_year_cache.get(int(year)) or session.execute(select(CampYear).where(CampYear.year == int(year))).scalar_one_or_none()
        if camp_year is None:
            camp_year = CampYear(year=int(year), name=f"Zeltlager {int(year)}")
            session.add(camp_year)
            counters.camp_years += 1
            camp_year_cache[int(year)] = camp_year

        recipe = recipe_cache.get(normalize_name(str(recipe_name))) or session.execute(
            select(Recipe).where(Recipe.normalized_name == normalize_name(str(recipe_name)))
        ).scalar_one_or_none()
        if recipe is None:
            issues.append(ImportIssueRecord("warning", ws.title, f"B{row}", "Feedback-Rezept wurde nicht gefunden.", str(recipe_name)))
            continue

        # Ein Feedback gilt pro (Camp-Jahr, Rezept), nicht mehr je Mahlzeit-Slot - eine zweite Zeile
        # fuer dasselbe Jahr/Rezept in der Excel-Tabelle wuerde sonst am Unique-Constraint scheitern.
        session.flush()
        existing_feedback = session.execute(
            select(RecipeFeedback).where(
                RecipeFeedback.camp_year_id == camp_year.id, RecipeFeedback.recipe_id == recipe.id
            )
        ).scalar_one_or_none()
        if existing_feedback is not None:
            issues.append(
                ImportIssueRecord(
                    "warning",
                    ws.title,
                    f"A{row}",
                    "Weiteres Feedback fuer dieselbe Jahr/Rezept-Kombination gefunden - uebersprungen "
                    "(ein Feedback gilt jetzt pro Rezept und Jahr, nicht mehr je Mahlzeit-Slot).",
                    f"{recipe_name} ({year})",
                )
            )
            continue

        repeat_value = ws.cell(row=row, column=4).value
        repeat_next_time = None
        if isinstance(repeat_value, str):
            repeat_next_time = repeat_value.strip().lower() in {"ja", "yes", "true", "1"}

        session.add(
            RecipeFeedback(
                camp_year=camp_year,
                recipe=recipe,
                rating=int(ws.cell(row=row, column=3).value) if isinstance(ws.cell(row=row, column=3).value, (int, float)) else None,
                repeat_next_time=repeat_next_time,
                planned_portions=int(ws.cell(row=row, column=5).value) if isinstance(ws.cell(row=row, column=5).value, (int, float)) else None,
                cooked_portions=int(ws.cell(row=row, column=6).value) if isinstance(ws.cell(row=row, column=6).value, (int, float)) else None,
                leftover_quantity=parse_decimal(ws.cell(row=row, column=7).value),
                leftover_unit=normalize_unit(ws.cell(row=row, column=8).value),
                quantity_factor_next_time=parse_decimal(ws.cell(row=row, column=9).value),
                process_tips=str(ws.cell(row=row, column=10).value) if ws.cell(row=row, column=10).value else None,
                what_went_well=str(ws.cell(row=row, column=11).value) if ws.cell(row=row, column=11).value else None,
                what_to_change=str(ws.cell(row=row, column=12).value) if ws.cell(row=row, column=12).value else None,
            )
        )
        counters.recipe_feedback += 1


def import_shopping_sheet(ws: Worksheet, session, ingredient_cache: dict[str, Ingredient], camp_year: CampYear | None, counters: ImportCounters, issues: list[ImportIssueRecord]) -> None:
    if ws.cell(row=1, column=1).value == "Zutat":
        header_row = 1
        start_row = 2
    else:
        header_row = 6
        start_row = 7

    shopping_list = ShoppingList(
        camp_year=camp_year if camp_year is not None else session.execute(select(CampYear).where(CampYear.year == 2026)).scalar_one_or_none(),
        name=ws.title,
        notes=f"Importiert aus {ws.title}",
    )
    session.add(shopping_list)
    counters.shopping_lists += 1

    for row in range(start_row, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        quantity = ws.cell(row=row, column=2).value
        unit = ws.cell(row=row, column=3).value
        normalized_unit = normalize_unit(unit)
        if should_skip_shopping_row(name, normalized_unit):
            continue

        decimal_quantity = parse_decimal(quantity)
        if decimal_quantity is None:
            issues.append(ImportIssueRecord("warning", ws.title, f"B{row}", "Einkaufsmenge konnte nicht gelesen werden.", str(quantity)))
            continue

        ingredient, created = resolve_or_create_ingredient(session, ingredient_cache, str(name), normalized_unit)
        if created:
            counters.ingredients += 1

        shopping_list.items.append(
            ShoppingListItem(
                ingredient=ingredient,
                quantity=decimal_quantity,
                unit=normalized_unit,
                estimated_price_per_unit=parse_decimal(ws.cell(row=row, column=4).value),
                estimated_total_price=parse_decimal(ws.cell(row=row, column=7).value if header_row == 6 else ws.cell(row=row, column=5).value),
                shopping_date=parse_excel_date(ws.cell(row=row, column=8).value) if header_row == 6 else None,
                status=str(ws.cell(row=row, column=10).value if header_row == 6 else None) if (header_row == 6 and ws.cell(row=row, column=10).value) else None,
                linked_recipes_text=str(ws.cell(row=row, column=6).value if header_row == 6 else ws.cell(row=row, column=6).value) if ws.cell(row=row, column=6).value else None,
                notes=str(ws.cell(row=row, column=11).value) if header_row == 6 and ws.cell(row=row, column=11).value else None,
            )
        )
        counters.shopping_list_items += 1


def persist_issues(session, import_run: ImportRun, issues: list[ImportIssueRecord], counters: ImportCounters) -> None:
    for issue in issues:
        session.add(
            ImportIssue(
                import_run=import_run,
                severity=issue.severity,
                sheet_name=issue.sheet_name,
                cell_reference=issue.cell_reference,
                message=issue.message,
                raw_value=issue.raw_value,
            )
        )
        counters.import_issues += 1


def write_import_report(project_root: Path, source_file: Path, counters: ImportCounters, issues: list[ImportIssueRecord], database_path: Path) -> None:
    docs_dir = project_root / "docs"
    docs_dir.mkdir(parents=True, exist_ok=True)
    payload = {
        "source_file": str(source_file),
        "database_path": str(database_path),
        "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "counters": asdict(counters),
        "issues": [asdict(issue) for issue in issues],
    }
    (docs_dir / "import_run_report.json").write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")

    lines = [
        "# Import-Report",
        "",
        f"- Quelle: `{source_file}`",
        f"- Ziel-Datenbank: `{database_path}`",
        "",
        "## Importierte Datensaetze",
        "",
    ]
    for key, value in asdict(counters).items():
        lines.append(f"- `{key}`: `{value}`")
    lines.extend(["", "## Import-Issues", ""])
    if not issues:
        lines.append("- Keine Import-Issues protokolliert")
    else:
        for issue in issues:
            location = f"{issue.sheet_name or '-'} {issue.cell_reference or ''}".strip()
            lines.append(f"- `{issue.severity}` in `{location}`: {issue.message} {issue.raw_value or ''}".rstrip())
    (docs_dir / "import_run_report.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def run_import(excel_path: Path, config: AppConfig) -> tuple[ImportCounters, list[ImportIssueRecord]]:
    values_wb = load_workbook(excel_path, data_only=True)
    formulas_wb = load_workbook(excel_path, data_only=False)
    _, _, session_factory = initialize_database(config)

    counters = ImportCounters()
    issues: list[ImportIssueRecord] = []

    with session_scope(session_factory) as session:
        import_run = ImportRun(source_file=str(excel_path), status="running", notes="Excel-Migration gestartet")
        session.add(import_run)

        ingredient_cache: dict[str, Ingredient] = {}
        recipe_cache: dict[str, Recipe] = {}
        camp_year_cache: dict[int, CampYear] = {}
        known_units = build_known_units_map(values_wb)

        for sheet_name in ["Preisliste", "Preisliste 2024"]:
            if sheet_name in formulas_wb.sheetnames:
                import_price_list(formulas_wb[sheet_name], session, ingredient_cache, counters, issues)

        for formula_ws in formulas_wb.worksheets:
            if is_recipe_sheet(formula_ws.title):
                import_recipe_sheet(
                    values_wb[formula_ws.title], formula_ws, session, ingredient_cache, recipe_cache,
                    known_units, counters, issues,
                )

        camp_year = None
        if "Planung 2026" in values_wb.sheetnames:
            camp_year = import_planning_sheet(values_wb["Planung 2026"], session, recipe_cache, counters, issues)
            if camp_year is not None:
                camp_year_cache[camp_year.year] = camp_year

        if "Rezept Feedback" in values_wb.sheetnames:
            import_feedback_sheet(values_wb["Rezept Feedback"], session, recipe_cache, camp_year_cache, counters, issues)

        for shopping_name in ["Einkaufsliste 2024", "Einkaufsliste 2025", "Einkaufsplanung"]:
            if shopping_name in values_wb.sheetnames:
                import_shopping_sheet(values_wb[shopping_name], session, ingredient_cache, camp_year, counters, issues)

        persist_issues(session, import_run, issues, counters)
        import_run.status = "completed"
        import_run.notes = f"Import abgeschlossen mit {counters.import_issues} Issues."

    write_import_report(config.project_root, excel_path, counters, issues, config.database_path)
    return counters, issues


def main() -> int:
    parser = argparse.ArgumentParser(description="Migrate the Excel workbook into the SQLite database.")
    parser.add_argument("--excel-path", help="Optional path to the workbook.")
    parser.add_argument("--db-path", help="Optional path to the SQLite database.")
    args = parser.parse_args()

    project_root = PROJECT_ROOT
    excel_path = find_excel_file(project_root, args.excel_path)
    config = AppConfig.load(project_root=project_root, database_path=Path(args.db_path) if args.db_path else None)

    counters, issues = run_import(excel_path, config)
    print(f"Quelle: {excel_path}")
    print(f"Ziel-Datenbank: {config.database_path}")
    for key, value in asdict(counters).items():
        print(f"{key}: {value}")
    print(f"Issues: {len(issues)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
