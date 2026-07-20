from openpyxl import Workbook
from sqlalchemy import select

from app.config import AppConfig
from app.db import create_engine_from_config, create_session_factory, init_database, session_scope
from app.models import CampYear, Ingredient, Recipe
from scripts.migrate_excel_to_sqlite import run_import


def build_test_workbook(path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Preisliste"
    ws.append(["Zutat", "Preis", "Einheit", "Status", "Quelle / Shop", "Stand", "Preisnotiz"])
    ws.append(["Nudeln", 2.49, "€/kg", "OK", "Testmarkt", None, None])
    ws.append(["Tomaten", 3.10, "€/kg", "OK", "Testmarkt", None, None])

    recipe = wb.create_sheet("Testgericht")
    recipe["E2"] = "Testgericht"
    recipe["G6"] = "Portionen:"
    recipe["H6"] = 10
    recipe.append([])
    recipe.append(["Zutaten:", "Grundmenge:", "Einheit:", "Gesamtmenge", "Preis/kg:", "Gesamtpreis:"])
    recipe.append(["Nudeln", 0.15, "kg", 1.5, 2.49, 3.735])
    recipe.append(["Tomaten", 0.05, "kg", 0.5, 3.10, 1.55])
    recipe.append(["Schritte:", "Kochen", 20])
    recipe.append(["Gesamtdauer:", 20, "Minuten"])

    planning = wb.create_sheet("Planung 2026")
    planning["A4"] = 2026
    planning["B4"] = None
    planning["C4"] = None
    planning.append([])
    planning.append(["Datum", "Wochentag", "Mahlzeit", "Rezept", "Portionen", "Zielgruppe", "Einkaufstag", "Einkaufsgruppe", "Status", "Kosten geplant", "Einkauf erledigt?"])
    planning.append([None, None, "Abend", "Testgericht", 95, "Kinder", None, "Gruppe A", "geplant", 0, "offen"])

    feedback = wb.create_sheet("Rezept Feedback")
    feedback.append(["Rezeptbewertung & Erfahrungswissen pro Jahr"])
    feedback.append([])
    feedback.append(["Jahr", "Rezept", "Bewertung 1-5", "Wiederholen?", "Portionen geplant", "Portionen gekocht", "Übrig geblieben", "Einheit Rest", "Mengenfaktor nächstes Mal", "Ablauf-Tipps / Tricks", "Was lief gut?", "Was ändern?"])
    feedback.append([2026, "Testgericht", 5, "ja", 95, 100, 2, "kg", 1.05, "Schnell", "Gut", "Nichts"])

    shopping = wb.create_sheet("Einkaufsliste 2025")
    shopping.append(["Zutat", "Menge", "Einheit", "Preis pro Einheit", "Gesamtpreis", "Gerichte"])
    shopping.append(["Nudeln", 14.25, "kg", 2.49, 35.48, "Testgericht"])

    price_2024 = wb.create_sheet("Preisliste 2024")
    price_2024.append(["Zutat", "Preis", "Einheit"])
    price_2024.append(["Nudeln", 1.99, "€/kg"])

    wb.save(path)


def test_run_import_populates_database_and_reports(tmp_path) -> None:
    workbook_path = tmp_path / "test_import.xlsx"
    build_test_workbook(workbook_path)

    config = AppConfig.load(project_root=tmp_path, database_path=tmp_path / "instance" / "import.sqlite3")
    counters, issues = run_import(workbook_path, config)

    assert counters.recipes >= 1
    assert counters.recipe_ingredients >= 2
    assert counters.ingredient_prices >= 2
    assert len(issues) == 0
    assert (tmp_path / "docs" / "import_run_report.md").exists()
    assert (tmp_path / "docs" / "import_run_report.json").exists()

    engine = create_engine_from_config(config)
    init_database(engine)
    session_factory = create_session_factory(engine)
    with session_scope(session_factory) as session:
        assert session.execute(select(Recipe).where(Recipe.normalized_name == "testgericht")).scalar_one()
        assert session.execute(select(Ingredient).where(Ingredient.normalized_name == "nudeln")).scalar_one()
        assert session.execute(select(CampYear).where(CampYear.year == 2026)).scalar_one()


def test_run_import_is_idempotent_for_price_list_rows(tmp_path) -> None:
    """Re-running the import must not duplicate price-list rows (regression: no dedup check existed)."""
    workbook_path = tmp_path / "test_import.xlsx"
    build_test_workbook(workbook_path)
    config = AppConfig.load(project_root=tmp_path, database_path=tmp_path / "instance" / "import.sqlite3")

    run_import(workbook_path, config)
    second_counters, _ = run_import(workbook_path, config)

    assert second_counters.ingredient_prices == 0

    engine = create_engine_from_config(config)
    init_database(engine)
    session_factory = create_session_factory(engine)
    with session_scope(session_factory) as session:
        noodles = session.execute(select(Ingredient).where(Ingredient.normalized_name == "nudeln")).scalar_one()
        # Preisliste + Preisliste 2024 + the recipe sheet's own price line - not doubled by the second run.
        assert len(noodles.prices) == 3
