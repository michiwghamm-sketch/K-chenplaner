from decimal import Decimal

from openpyxl import Workbook
from sqlalchemy import select

from app.config import AppConfig
from app.db import create_engine_from_config, create_session_factory, init_database, session_scope
from app.models import Ingredient, Recipe, RecipeComponent, RecipeIngredient
from scripts.migrate_excel_to_sqlite import run_import


def build_workbook_with_components(path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Preisliste"
    ws.append(["Zutat", "Preis", "Einheit", "Status", "Quelle / Shop", "Stand", "Preisnotiz"])
    ws.append(["Paprika Pulver", 8.99, "kg", "OK", "Testmarkt", None, None])

    recipe = wb.create_sheet("Testrezept mit Teilen")
    recipe["E2"] = "Testrezept mit Teilen"
    recipe["G6"] = "Portionen:"
    recipe["H6"] = 20
    recipe.append([])
    recipe.append(["Zutaten:", "Grundmenge:", "Einheit:", "Gesamtmenge", "Preis/kg:", "Gesamtpreis:", "Teil A"])
    recipe.append(["Hackfleisch", 0.1, "kg", 2, 12, 24])
    recipe.append(["Zwiebel", 0.05, "kg", 1, 1.5, 1.5])
    recipe.append([])
    recipe.append(["Kartoffel", 0.2, "kg", 4, 1.5, 6, "Teil B"])
    # missing unit, has quantity -> should fall back to known unit "kg" from the Preisliste row.
    recipe.append(["Paprika Pulver geraeuchert", 0.002, None, 0.04, 8.99, 0.36])
    # missing quantity entirely ("to taste") -> should import as optional with quantity 0.
    recipe.append(["Muskat", None, None, 0, None, 0])
    recipe.append(["Gesamtkosten:", None, None, None, None, 31.86])
    recipe.append(["Zubereitung: "])
    recipe.append(["Schritte:", "Anweisung:", "Ungefaehre Dauer in min:"])
    recipe.append(["Anbraten", "Alles anbraten", 10])

    # Component labels live in merged G/H cells spanning the rows that belong to them.
    # Row layout: 8=header, 9-10=Teil A ingredients, 11=blank separator, 12-14=Teil B ingredients.
    recipe.merge_cells("G9:H10")
    recipe["G9"] = "Teil A"
    recipe.merge_cells("G12:H14")
    recipe["G12"] = "Teil B"

    flat_recipe = wb.create_sheet("Flaches Rezept")
    flat_recipe["E2"] = "Flaches Rezept"
    flat_recipe["H6"] = 10
    flat_recipe.append([])
    flat_recipe.append(["Zutaten:", "Grundmenge:", "Einheit:", "Gesamtmenge", "Preis/kg:", "Gesamtpreis:"])
    flat_recipe.append(["Reis", 0.1, "kg", 1, 2, 2])
    flat_recipe.append(["Gesamtkosten:", None, None, None, None, 2])
    # Whole-block label repeating the recipe's own name -> must NOT become a component.
    flat_recipe.merge_cells("G9:H11")
    flat_recipe["G9"] = "Flaches Rezept"

    # Regression case: label repeats the recipe name but without the space
    # (mirrors the real workbook's "Gemüse Nudeln" recipe labelled "Gemüsenudeln").
    spaceless_recipe = wb.create_sheet("Gemuese Nudeln")
    spaceless_recipe["E2"] = "Gemuese Nudeln"
    spaceless_recipe["H6"] = 10
    spaceless_recipe.append([])
    spaceless_recipe.append(["Zutaten:", "Grundmenge:", "Einheit:", "Gesamtmenge", "Preis/kg:", "Gesamtpreis:"])
    spaceless_recipe.append(["Nudeln", 0.1, "kg", 1, 2, 2])
    spaceless_recipe.append(["Gesamtkosten:", None, None, None, None, 2])
    spaceless_recipe.merge_cells("G9:H11")
    spaceless_recipe["G9"] = "GemueseNudeln"

    wb.save(path)


def test_migration_creates_components_and_assigns_ingredients(tmp_path) -> None:
    workbook_path = tmp_path / "components.xlsx"
    build_workbook_with_components(workbook_path)

    config = AppConfig.load(project_root=tmp_path, database_path=tmp_path / "instance" / "components.sqlite3")
    counters, issues = run_import(workbook_path, config)

    assert counters.recipe_components == 2

    engine = create_engine_from_config(config)
    init_database(engine)
    session_factory = create_session_factory(engine)
    with session_scope(session_factory) as session:
        recipe = session.execute(
            select(Recipe).where(Recipe.normalized_name == "testrezept mit teilen")
        ).scalar_one()
        components_by_name = {c.name: c for c in recipe.components}
        assert set(components_by_name) == {"Teil A", "Teil B"}

        ingredients_by_name = {item.ingredient.name: item for item in recipe.ingredients}
        assert ingredients_by_name["Hackfleisch"].component_id == components_by_name["Teil A"].id
        assert ingredients_by_name["Zwiebel"].component_id == components_by_name["Teil A"].id
        assert ingredients_by_name["Kartoffel"].component_id == components_by_name["Teil B"].id

        # Missing unit -> falls back to the unit known from the Preisliste ("kg").
        smoked_paprika = ingredients_by_name["Paprika Pulver geraeuchert"]
        assert smoked_paprika.unit == "kg"
        assert smoked_paprika.quantity == Decimal("0.002")
        assert smoked_paprika.component_id == components_by_name["Teil B"].id

        # Missing quantity entirely -> imported as optional, quantity 0, with a note.
        nutmeg = ingredients_by_name["Muskat"]
        assert nutmeg.quantity == Decimal("0")
        assert nutmeg.optional is True
        assert nutmeg.notes == "Menge nicht in Excel angegeben (nach Geschmack)"

    info_messages = [issue.message for issue in issues if issue.severity == "info"]
    assert any("Keine Einheit angegeben" in message for message in info_messages)
    assert any("Keine Menge angegeben" in message for message in info_messages)


def test_migration_does_not_create_component_for_whole_block_label(tmp_path) -> None:
    workbook_path = tmp_path / "flat.xlsx"
    build_workbook_with_components(workbook_path)

    config = AppConfig.load(project_root=tmp_path, database_path=tmp_path / "instance" / "flat.sqlite3")
    run_import(workbook_path, config)

    engine = create_engine_from_config(config)
    init_database(engine)
    session_factory = create_session_factory(engine)
    with session_scope(session_factory) as session:
        recipe = session.execute(select(Recipe).where(Recipe.normalized_name == "flaches rezept")).scalar_one()
        assert recipe.components == []
        assert recipe.ingredients[0].component_id is None


def test_migration_ignores_whole_block_label_missing_a_space(tmp_path) -> None:
    workbook_path = tmp_path / "flat.xlsx"
    build_workbook_with_components(workbook_path)

    config = AppConfig.load(project_root=tmp_path, database_path=tmp_path / "instance" / "flat2.sqlite3")
    run_import(workbook_path, config)

    engine = create_engine_from_config(config)
    init_database(engine)
    session_factory = create_session_factory(engine)
    with session_scope(session_factory) as session:
        recipe = session.execute(select(Recipe).where(Recipe.normalized_name == "gemuese nudeln")).scalar_one()
        assert recipe.components == []
        assert recipe.ingredients[0].component_id is None


def test_migration_is_idempotent_for_components_and_ingredients(tmp_path) -> None:
    workbook_path = tmp_path / "components.xlsx"
    build_workbook_with_components(workbook_path)
    config = AppConfig.load(project_root=tmp_path, database_path=tmp_path / "instance" / "components.sqlite3")

    run_import(workbook_path, config)
    counters_second_run, _ = run_import(workbook_path, config)

    # Second run must not create duplicate components/ingredients for data already imported.
    assert counters_second_run.recipe_components == 0
    assert counters_second_run.recipe_ingredients == 0

    engine = create_engine_from_config(config)
    init_database(engine)
    session_factory = create_session_factory(engine)
    with session_scope(session_factory) as session:
        recipe = session.execute(
            select(Recipe).where(Recipe.normalized_name == "testrezept mit teilen")
        ).scalar_one()
        assert len(recipe.components) == 2
        assert len(recipe.ingredients) == 5
